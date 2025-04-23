import os
import openpyxl
from datetime import datetime
import re
from systemLogger import Logger
logger = Logger()

PATH_OF_DATA = 'data'

# 從 firebase Storage 取得檔案
def download_file_from_storage(bucket, file_path):

  msg_source = "subjectInfo.py line.10"
  logger.debug(f"Attempting to download file from storage: {file_path}", msg_source)

  # 取得下載目標 blob
  blob = bucket.blob(file_path)

  # 本地存放檔案路徑
  local_file_path = os.path.join(PATH_OF_DATA, file_path)

  # 如果本地存放檔案路徑不存在，則建立目錄
  if not os.path.exists(os.path.dirname(local_file_path)):
    logger.debug(f"Creating directory: {os.path.dirname(local_file_path)}", msg_source)
    os.makedirs(os.path.dirname(local_file_path))

  # 下載檔案
  if not os.path.exists(local_file_path):
    blob.download_to_filename(local_file_path)

  logger.debug(f"File downloaded successfully: {local_file_path}", msg_source)

  # 回傳本地存放檔案路徑
  return local_file_path

def parse_subject_info(subject_file):

    gender_defined = {
        "Female": "female",
        "Male": "male",
        "男": "male",
        "女": "female",
    }

    type_defined = {
        "Blood": "blood",
        "Chorionic villi": "chorionicVilli",
        "Amniotic fluid": "amnioticFluid",
        "Cord blood": "cordBlood",
        "DBS": "dbs",
        "Buccal swab": "buccalSwab",
        "FFPE": "ffpe",
        "FFPE + Blood": "ffpeBlood",
        "RNA": "rna",
        "Others": "others",
    }

    try:
        workbook = openpyxl.load_workbook(subject_file)
        worksheet = workbook['Subject Info']
    except Exception as e:
        print(f"Error reading file: {e}")
        raise

    subject_info = {}
    for idx, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=1):
        sample_id, name, birth, gender, id_number, type_, collecting_date, received_date = row

        subject_info[idx] = {
            "sampleId": sample_id,
            "name": str(name) if name else "",
            "birth": datetime.strftime(birth, "%Y/%m/%d") if birth else "",
            "gender": gender_defined.get(gender, "") if gender else "",
            "idNumber": str(id_number) if id_number else "",
            "type": type_defined.get(type_, "") if type_ else "",
            "collectingDate": datetime.strftime(collecting_date, "%Y/%m/%d") if collecting_date else "",
            "receivedDate": datetime.strftime(received_date, "%Y/%m/%d") if received_date else "",
        }

    return subject_info

def parse_input_analysis_lims(subject_file):
    """
    從 Excel 檔案中提取病患資料並進行格式化處理

    參數:
        subject_file: Excel 檔案路徑或二進制數據

    返回:
        dict: 格式化後的病患資料字典
    """
    gender_defined = {
        '女': "female",
        '男': "male"
    }

    def date_transfer(date_str):
        """將民國年日期轉換為西元年日期"""
        if not date_str or not isinstance(date_str, str):
            return None

        parts = date_str.split('/')
        if len(parts) != 3:
            return None

        try:
            year = int(parts[0]) + 1911
            month = int(parts[1])
            day = int(parts[2])
            return datetime(year, month, day)
        except (ValueError, TypeError):
            return None

    try:
        # 讀取 Excel 檔案
        workbook = openpyxl.load_workbook(subject_file)
        worksheet = workbook['工作表1']

        # 獲取標題行
        headers = {}
        for col_idx, cell in enumerate(worksheet[1], 1):
            if cell.value:
                headers[cell.value] = col_idx

        # 檢查必要欄位是否存在
        required_columns = ['開單日期', '流水號', '姓名', '身分證號', '生日', '性別', '原就醫日']
        for col in required_columns:
            if col not in headers:
                raise ValueError(f"缺少必要欄位: {col}")

        # 讀取資料
        subject_info_all = []
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2), 2):
            row_data = {}

            # 讀取每個欄位的值
            for col_name, col_idx in headers.items():
                cell_value = row[col_idx-1].value
                row_data[col_name] = cell_value

            # 處理日期格式
            for date_field in ['開單日期', '生日', '原就醫日']:
                if date_field in row_data:
                    if isinstance(row_data[date_field], datetime):
                        row_data[date_field] = row_data[date_field].strftime('%Y/%m/%d')

            # 添加到列表
            subject_info_all.append({
                'orderingDate': row_data.get('開單日期', ''),
                'serialId': row_data.get('流水號', ''),
                'name': row_data.get('姓名', ''),
                'dob': row_data.get('生日', ''),
                'gender': row_data.get('性別', ''),
                'idNumber': row_data.get('身分證號', ''),
                'collectionDate': row_data.get('原就醫日', '')
            })

        # 處理數據轉換
        subject_info = {}
        for s in subject_info_all:
            # 跳過空行
            if not s['orderingDate'] or not s['serialId']:
                continue

            # 生成 ID
            id_str = str(s['serialId'])

            # 轉換資料格式
            subject_info[id_str] = {
                'name': str(s['name']) if s['name'] else "",
                'birth': date_transfer(s['dob']).strftime('%Y/%m/%d') if s['dob'] and date_transfer(s['dob']) else "",
                'gender': gender_defined.get(s['gender'], "") if s['gender'] else "",
                'type': "",
                'idNumber': str(s['idNumber']) if s['idNumber'] else "",
                'collectingDate': date_transfer(s['collectionDate']).strftime('%Y/%m/%d') if s['collectionDate'] and date_transfer(s['collectionDate']) else "",
                'receivedDate': date_transfer(s['orderingDate']).strftime('%Y/%m/%d') if s['orderingDate'] and date_transfer(s['orderingDate']) else "",
                'edit': True
            }

        return subject_info

    except Exception as e:
        print(f"處理 Excel 檔案時發生錯誤: {str(e)}")
        raise e